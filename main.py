import os
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import sqlite3
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import Toplevel

# Initialize or load Excel files
def create_or_load_excel_files():
    deposit_filename = "deposit_entries.xlsx"
    stock_option_filename = "stock_option_entries.xlsx"

    if not os.path.exists(deposit_filename):
        deposit_workbook = openpyxl.Workbook()
        deposit_sheet = deposit_workbook.active
        deposit_sheet.title = 'Deposit Entries'
        deposit_sheet.append(["Client ID", "Client Name", "Deposited Amount", "Balance", "Date", "Time"])
        deposit_workbook.save(deposit_filename)

    if not os.path.exists(stock_option_filename):
        stock_option_workbook = openpyxl.Workbook()
        stock_option_sheet = stock_option_workbook.active
        stock_option_sheet.title = 'Stock Option Entries'
        stock_option_sheet.append(["Client ID", "Client Name", "Stock/Option", "Strike/Call", "CE/PE", 
                                   "Lots/Quantity", "Amount in INR", "P&L in INR", "P&L in %", "Date", "Time"])
        stock_option_workbook.save(stock_option_filename)

def create_tables():
    conn = sqlite3.connect("trading_entry_book.db")
    cursor = conn.cursor()

    # Create deposits table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS deposits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id TEXT,
        client_name TEXT,
        deposited_amount REAL,
        balance REAL,
        date TEXT,
        time TEXT
    )
    """)

    # Create stock/option table
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS stock_options (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id TEXT,
        client_name TEXT,
        stock_option_type TEXT,
        strike_call TEXT,
        ce_pe TEXT,
        lots_qty INTEGER,
        amount_inr REAL,
        pl_inr REAL,
        pl_percentage REAL,
        date TEXT,
        time TEXT
    )
    """)

    conn.commit()
    conn.close()

def create_or_load_sqldb():
    # Connect to the SQLite database
    conn = sqlite3.connect("trading_entry_book.db")
    cursor = conn.cursor()

    # Call the function to create tables if they don't exist
    create_tables()  # Make sure this line is present

    conn.commit()
    conn.close()



# Save deposit entry to Excel
def save_deposit_entry1(client_id, client_name, deposited_amount, balance, date_str, time_str):
    deposit_filename = "deposit_entries.xlsx"
    try:
        workbook = openpyxl.load_workbook(deposit_filename)
        sheet = workbook.active
        sheet.append([client_id, client_name, deposited_amount, balance, date_str, time_str])
        workbook.save(deposit_filename)
        messagebox.showinfo("Success", "Deposit entry saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save entry: {e}")

def save_deposit_entry(client_id, client_name, deposited_amount, balance, date_str, time_str):
    conn = sqlite3.connect('trading_entry_book.db')  # Ensure database filename matches
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO deposits (client_id, client_name, deposited_amount, balance, date, time) VALUES (?, ?, ?, ?, ?, ?) ",
                       (client_id, client_name, deposited_amount, balance, date_str, time_str))
        conn.commit()
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"Failed to save deposit entry: {e}")
    finally:
        conn.close()

# Save stock/option entry to Excel
def save_stock_option1(client_id, client_name, stock_option_type, strike_call, ce_pe, lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str):
    stock_option_filename = "stock_option_entries.xlsx"
    try:
        workbook = openpyxl.load_workbook(stock_option_filename)
        sheet = workbook.active
        sheet.append([client_id, client_name, stock_option_type, strike_call, ce_pe, lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str])
        workbook.save(stock_option_filename)
        messagebox.showinfo("Success", f"{stock_option_type} entry saved successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save entry: {e}")

def save_stock_option(client_id, client_name, stock_option_type, strike_call, ce_pe, lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str):
    conn = sqlite3.connect('trading_entry_book.db')  # Ensure database filename matches
    cursor = conn.cursor()
    try:
        cursor.execute('''INSERT INTO stock_options
                        (client_id, client_name, stock_option_type, strike_call, ce_pe, lots_qty, amount_inr, pl_inr, pl_percentage, date, time)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                       (client_id, client_name, stock_option_type, strike_call, ce_pe, lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str))
        conn.commit()
    except sqlite3.Error as e:
        messagebox.showerror("Error", f"Failed to save stock/option entry: {e}")
    finally:
        conn.close()

# Function to check if a client exists in the deposits table
def client_exists(client_id, client_name):
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM deposits WHERE client_id = ? AND client_name = ?", (client_id, client_name))
    exists = cursor.fetchone() is not None
    conn.close()
    return exists

def plot_deposit_distribution():
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()
    cursor.execute("SELECT client_name, SUM(deposited_amount) FROM deposits GROUP BY client_name")
    data = cursor.fetchall()
    conn.close()

    if data:
        client_names, amounts = zip(*data)

        plt.figure(figsize=(8, 8))
        plt.pie(amounts, labels=client_names, autopct='%1.1f%%')
        plt.title('Deposit Distribution by Client')
        plt.axis('equal')
        plt.show()
    else:
        messagebox.showinfo("Info", "No data available for plotting.")

# Function to calculate P&L for a given date range
def calculate_portfolio_performance(start_date, end_date):
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()

    # Query to get P&L for each client
    cursor.execute("""
        SELECT client_name, SUM(pl_inr), SUM(amount_inr) 
        FROM stock_options 
        WHERE date BETWEEN ? AND ? 
        GROUP BY client_name
    """, (start_date, end_date))
    
    results = cursor.fetchall()
    conn.close()

    overall_pl_inr = sum(row[1] for row in results)
    overall_amount_inr = sum(row[2] for row in results)

    return results, overall_pl_inr, overall_amount_inr

# Function to show the portfolio analysis window
def show_portfolio_analysis():
    analysis_window = Toplevel(root)
    analysis_window.title("Portfolio Analysis")

    # Date selection
    tk.Label(analysis_window, text="Start Date (YYYY-MM-DD):").grid(row=0, column=0)
    start_date_entry = tk.Entry(analysis_window)
    start_date_entry.grid(row=0, column=1)

    tk.Label(analysis_window, text="End Date (YYYY-MM-DD):").grid(row=1, column=0)
    end_date_entry = tk.Entry(analysis_window)
    end_date_entry.grid(row=1, column=1)

    def analyze_portfolio():
        start_date = start_date_entry.get()
        end_date = end_date_entry.get()

        # Validate dates
        try:
            start_date_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_date_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Please enter valid dates in YYYY-MM-DD format.")
            return

        if end_date_dt < start_date_dt:
            messagebox.showerror("Error", "End date must be after start date.")
            return

        results, overall_pl_inr, overall_amount_inr = calculate_portfolio_performance(start_date, end_date)

        # Show results in a new window
        result_window = Toplevel(analysis_window)
        result_window.title("Portfolio Performance Results")

        # Overall P&L
        tk.Label(result_window, text="Overall P&L in INR:").grid(row=0, column=0)
        tk.Label(result_window, text=overall_pl_inr).grid(row=0, column=1)

        tk.Label(result_window, text="Overall Amount in INR:").grid(row=1, column=0)
        tk.Label(result_window, text=overall_amount_inr).grid(row=1, column=1)

        tk.Label(result_window, text="Client-wise Performance:").grid(row=2, column=0, columnspan=2)

        for index, (client_name, pl_inr, amount_inr) in enumerate(results, start=3):
            tk.Label(result_window, text=client_name).grid(row=index, column=0)
            tk.Label(result_window, text=f"P&L: {pl_inr} INR").grid(row=index, column=1)
            tk.Label(result_window, text=f"Amount: {amount_inr} INR").grid(row=index, column=2)

    analyze_button = tk.Button(analysis_window, text="Analyze", command=analyze_portfolio)
    analyze_button.grid(row=2, column=0, columnspan=2, pady=10)


def analyze_deposits():
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()
    cursor.execute("SELECT SUM(deposited_amount), COUNT(*) FROM deposits")
    total_deposits, deposit_count = cursor.fetchone()
    avg_deposit = total_deposits / deposit_count if deposit_count > 0 else 0
    conn.close()

    messagebox.showinfo("Deposit Analysis", 
                        f"Total Deposits: {total_deposits}\n"
                        f"Average Deposit: {avg_deposit:.2f}\n"
                        f"Total Entries: {deposit_count}")

# Function to get current date/time
def get_current_datetime():
    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d')
    time_str = now.strftime('%H:%M:%S')
    return date_str, time_str

# Validation function to check if fields are empty
def validate_entries(*entries):
    return all(entry.get().strip() for entry in entries)

# Reset fields after entry is saved
def reset_entries(*entries):
    for entry in entries:
        entry.delete(0, tk.END)

# Update Treeview for deposits
def update_deposit_treeview(treeview):
    deposit_filename = "deposit_entries.xlsx"
    workbook = openpyxl.load_workbook(deposit_filename)
    sheet = workbook.active
    treeview.delete(*treeview.get_children())

    for row in sheet.iter_rows(min_row=2, values_only=True):
        client_id = row[0]
        values = row[1:]

        # Insert values into the treeview
        treeview.insert('', 'end', values=(client_id, *values))

# Update Treeview for stock/option entries
def update_stock_option_treeview(treeview):
    stock_option_filename = "stock_option_entries.xlsx"
    workbook = openpyxl.load_workbook(stock_option_filename)
    sheet = workbook.active
    treeview.delete(*treeview.get_children())
    for row in sheet.iter_rows(min_row=2, values_only=True):
        treeview.insert('', 'end', values=row)

# Create Treeview for deposit entries with delete button
def create_deposit_treeview(frame):
    columns = ("Client ID", "Client Name", "Deposited Amount", "Balance", "Date", "Time")
    treeview = ttk.Treeview(frame, columns=columns, show="headings", height=10)

    for col in columns:
        treeview.heading(col, text=col)
        treeview.column(col, width=120)

    treeview.pack(fill=tk.BOTH, expand=True)

    scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=treeview.yview)
    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    treeview.config(yscrollcommand=scrollbar_y.set)

    return treeview

# Create Treeview for stock/option entries
def create_stock_option_treeview(frame):
    columns = ("Client ID", "Client Name", "Stock/Option", "Strike/Call", "CE/PE", "Lots/Quantity", 
               "Amount in INR", "P&L in INR", "P&L in %", "Date", "Time")
    treeview = ttk.Treeview(frame, columns=columns, show="headings", height=10)
    for col in columns:
        treeview.heading(col, text=col)
        treeview.column(col, width=100)
    
    treeview.pack(fill=tk.BOTH, expand=True)
    
    scrollbar_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=treeview.yview)
    scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
    treeview.config(yscrollcommand=scrollbar_y.set)
    
    return treeview

# Update Treeview with real-time data
def update_real_time_data():
    # Fetch data from the database
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM deposits")
    deposits = cursor.fetchall()

    cursor.execute("SELECT * FROM stock_options")
    stock_options = cursor.fetchall()
    conn.close()

    # Clear the treeview
    for i in deposit_tree.get_children():
        deposit_tree.delete(i)
    for i in stock_tree.get_children():
        stock_tree.delete(i)

    # Insert fresh data
    for deposit in deposits:
        deposit_tree.insert("", "end", values=deposit)

    for stock_option in stock_options:
        stock_tree.insert("", "end", values=stock_option)

    # Schedule next update
    root.after(5000, update_real_time_data)  # Update every 5 seconds



# Function to plot portfolio performance
def plot_portfolio_performance():
    conn = sqlite3.connect('trading_entry_book.db')
    cursor = conn.cursor()
    cursor.execute("SELECT date, SUM(deposited_amount) FROM deposits GROUP BY date")
    data = cursor.fetchall()
    conn.close()

    if data:
        dates, amounts = zip(*data)

        plt.figure(figsize=(10, 5))
        plt.plot(dates, amounts, marker='o')
        plt.title('Portfolio Performance Over Time')
        plt.xlabel('Date')
        plt.ylabel('Total Amount')
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Show the plot in a new window
        window = Toplevel(root)
        window.title("Portfolio Performance")
        canvas = FigureCanvasTkAgg(plt.gcf(), master=window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    else:
        messagebox.showinfo("Info", "No data available for plotting.")



# Function to delete a deposit entry
def delete_selected_deposit_entry(treeview):
    selected_item = treeview.selection()
    if not selected_item:
        messagebox.showerror("Error", "Please select a client to delete.")
        return

    client_id = treeview.item(selected_item, 'values')[0]
    
    if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete client ID {client_id}?"):
        conn = sqlite3.connect('trading_entry_book.db')
        cursor = conn.cursor()

        try:
            # Delete from deposits table
            cursor.execute("DELETE FROM deposits WHERE id = ?", (client_id,))
            print(f"Deleted from deposits where client_id = {client_id}")

            # Delete from stock_options table
            cursor.execute("DELETE FROM stock_options WHERE id = ?", (client_id,))
            print(f"Deleted from stock_options where client_id = {client_id}")

            conn.commit()
            messagebox.showinfo("Success", "Client deleted successfully.")
            update_deposit_treeview(treeview)  # Refresh treeview
            update_stock_option_treeview(stock_tree)  # Refresh stock treeview
        except sqlite3.Error as e:
            messagebox.showerror("Error", f"Failed to delete client: {e}")
        finally:
            conn.close()



# Deposit entry window
def deposit_entry_window(treeview):
    deposit_window = Toplevel(root)
    deposit_window.title("Add Deposit Entry")

    tk.Label(deposit_window, text="Client ID").grid(row=0, column=0)
    client_id_entry = tk.Entry(deposit_window)
    client_id_entry.grid(row=0, column=1)

    tk.Label(deposit_window, text="Client Name").grid(row=1, column=0)
    client_name_entry = tk.Entry(deposit_window)
    client_name_entry.grid(row=1, column=1)

    tk.Label(deposit_window, text="Deposited Amount").grid(row=2, column=0)
    deposited_amount_entry = tk.Entry(deposit_window)
    deposited_amount_entry.grid(row=2, column=1)

    tk.Label(deposit_window, text="Balance").grid(row=3, column=0)
    balance_entry = tk.Entry(deposit_window)
    balance_entry.grid(row=3, column=1)

    date_str, time_str = get_current_datetime()

    def save_deposit_action():
        client_id = client_id_entry.get()
        client_name = client_name_entry.get()
        deposited_amount = deposited_amount_entry.get()
        balance = balance_entry.get()
        
        if validate_entries(client_id_entry, client_name_entry, deposited_amount_entry, balance_entry):
            save_deposit_entry1(client_id, client_name, deposited_amount, balance, date_str, time_str)
            save_deposit_entry(client_id, client_name, deposited_amount, balance, date_str, time_str)
            update_deposit_treeview(treeview)
            reset_entries(client_id_entry, client_name_entry, deposited_amount_entry, balance_entry)
            deposit_window.destroy()  # Close the window after saving
        else:
            messagebox.showerror("Error", "Please fill in all fields")

    submit_btn = tk.Button(deposit_window, text="Save", command=save_deposit_action)
    submit_btn.grid(row=4, column=0, columnspan=2, pady=10)

# Stock/Option entry window
def stock_option_entry_window(treeview):
    stock_option_window = Toplevel(root)
    stock_option_window.title("Add Stock/Option Entry")

    tk.Label(stock_option_window, text="Client ID").grid(row=0, column=0)
    client_id_entry = tk.Entry(stock_option_window)
    client_id_entry.grid(row=0, column=1)

    tk.Label(stock_option_window, text="Client Name").grid(row=1, column=0)
    client_name_entry = tk.Entry(stock_option_window)
    client_name_entry.grid(row=1, column=1)

    tk.Label(stock_option_window, text="Stock/Option").grid(row=2, column=0)
    stock_option_type_entry = tk.Entry(stock_option_window)
    stock_option_type_entry.grid(row=2, column=1)

    tk.Label(stock_option_window, text="Strike/Call").grid(row=3, column=0)
    strike_call_entry = tk.Entry(stock_option_window)
    strike_call_entry.grid(row=3, column=1)

    tk.Label(stock_option_window, text="CE/PE").grid(row=4, column=0)
    ce_pe_entry = tk.Entry(stock_option_window)
    ce_pe_entry.grid(row=4, column=1)

    tk.Label(stock_option_window, text="Lots/Quantity").grid(row=5, column=0)
    lots_qty_entry = tk.Entry(stock_option_window)
    lots_qty_entry.grid(row=5, column=1)

    tk.Label(stock_option_window, text="Amount in INR").grid(row=6, column=0)
    amount_inr_entry = tk.Entry(stock_option_window)
    amount_inr_entry.grid(row=6, column=1)

    tk.Label(stock_option_window, text="P&L in INR").grid(row=7, column=0)
    pl_inr_entry = tk.Entry(stock_option_window)
    pl_inr_entry.grid(row=7, column=1)

    tk.Label(stock_option_window, text="P&L in %").grid(row=8, column=0)
    pl_percentage_entry = tk.Entry(stock_option_window)
    pl_percentage_entry.grid(row=8, column=1)

    date_str, time_str = get_current_datetime()

    def save_stock_option_action():
        client_id = client_id_entry.get()
        client_name = client_name_entry.get()
        stock_option_type = stock_option_type_entry.get()
        strike_call = strike_call_entry.get()
        ce_pe = ce_pe_entry.get()
        lots_qty = lots_qty_entry.get()
        amount_inr = amount_inr_entry.get()
        pl_inr = pl_inr_entry.get()
        pl_percentage = pl_percentage_entry.get()

        # Validate client ID and name
        if not client_exists(client_id, client_name):
            messagebox.showerror("Error", "Client ID and Name do not match any deposit entry.")
            return

        if validate_entries(client_id_entry, client_name_entry, stock_option_type_entry, strike_call_entry,
                            ce_pe_entry, lots_qty_entry, amount_inr_entry, pl_inr_entry, pl_percentage_entry):
            save_stock_option1(client_id, client_name, stock_option_type, strike_call, ce_pe,
                                lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str)
            save_stock_option(client_id, client_name, stock_option_type, strike_call, ce_pe,
                              lots_qty, amount_inr, pl_inr, pl_percentage, date_str, time_str)
            update_stock_option_treeview(treeview)
            reset_entries(client_id_entry, client_name_entry, stock_option_type_entry, strike_call_entry,
                           ce_pe_entry, lots_qty_entry, amount_inr_entry, pl_inr_entry, pl_percentage_entry)
            stock_option_window.destroy()  # Close the window after saving
        else:
            messagebox.showerror("Error", "Please fill in all fields")

    submit_btn = tk.Button(stock_option_window, text="Save", command=save_stock_option_action)
    submit_btn.grid(row=9, column=0, columnspan=2, pady=10)

# Main application
def main():
    global root, deposit_tree, stock_tree
    root = tk.Tk()
    root.title("Trading Entry Book")
    root.geometry("800x600")

    create_or_load_excel_files()
    create_tables()
    create_or_load_sqldb()

    deposit_frame = tk.Frame(root)
    deposit_frame.pack(fill=tk.BOTH, expand=True)
    deposit_tree = create_deposit_treeview(deposit_frame)

    stock_frame = tk.Frame(root)
    stock_frame.pack(fill=tk.BOTH, expand=True)
    stock_tree = create_stock_option_treeview(stock_frame)

    deposit_button = tk.Button(root, text="Add Deposit", command=lambda: deposit_entry_window(deposit_tree))
    deposit_button.pack(side=tk.LEFT, padx=10, pady=10)

    stock_option_button = tk.Button(root, text="Add Stock/Option", command=lambda: stock_option_entry_window(stock_tree))
    stock_option_button.pack(side=tk.LEFT, padx=10, pady=10)

    delete_button = tk.Button(root, text="Delete Client", command=lambda: delete_selected_deposit_entry(deposit_tree))
    delete_button.pack(side=tk.LEFT, padx=10, pady=10)

    analytics_button = tk.Button(root, text="Analyze Deposits", command=analyze_deposits)
    analytics_button.pack(side=tk.LEFT, padx=10, pady=10)

    distribution_button = tk.Button(root, text="Plot Deposit Distribution", command=plot_deposit_distribution)
    distribution_button.pack(side=tk.LEFT, padx=10, pady=10)

    portfolio_analysis_button = tk.Button(root, text="Portfolio Analysis", command=show_portfolio_analysis)
    portfolio_analysis_button.pack(side=tk.LEFT, padx=10, pady=10)

    plot_button = tk.Button(root, text="Plot Portfolio Performance", command=plot_portfolio_performance)
    plot_button.pack(side=tk.LEFT, padx=10, pady=10)

    update_deposit_treeview(deposit_tree)
    update_stock_option_treeview(stock_tree)

    update_real_time_data()

    root.mainloop()


# Run the main function
if __name__ == "__main__":
    main()


